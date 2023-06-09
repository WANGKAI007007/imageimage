import sys
import threading
import subprocess
from tkinter import filedialog, messagebox, scrolledtext
from openpyxl.styles import Font, PatternFill, Border, Side
from tkinter import ttk
from ttkthemes import ThemedTk
from datetime import datetime
import base64
import io
import logging
import os
import re
import requests
import tkinter as tk
from PIL import Image, ImageEnhance
from openpyxl import Workbook
import time


class TextHandler(logging.Handler):
    def __init__(self, text):
        logging.Handler.__init__(self)
        self.text = text

    def emit(self, record):
        msg = self.format(record)
        self.text.configure(state='normal')
        self.text.insert(tk.END, msg + '\n')
        self.text.configure(state='disabled')
        self.text.yview(tk.END)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s]: %(message)s",
    handlers=[
        logging.FileHandler("output.log"),
        logging.StreamHandler()
    ]
)


def select_directory(text_widget):
    directory = filedialog.askdirectory()
    if directory:
        text_widget.insert(tk.END, f'已选择文件夹 {directory}\n')
        text_widget.see(tk.END)
    return directory


def get_access_token(api_key, secret_key):
    token_url = "https://aip.baidubce.com/oauth/2.0/token"
    token_params = {
        "grant_type": "client_credentials",
        "client_id": api_key,
        "client_secret": secret_key,
    }
    try:
        response = requests.get(token_url, params=token_params)
        json_response = response.json()

        if 'access_token' in json_response:
            return json_response['access_token']
        else:
            logging.error("Unable to get access_token. Please check your API key and Secret key.")
            logging.error(json_response)
            return None
    except Exception as e:
        logging.error("Error occurred while getting access_token: %s", e)
        return None



def extract_invoice_fields(text):
    invoice_code_pattern = r'\d{10,12}'
    invoice_number_pattern = r'\d{8}'

    invoice_code = re.findall(invoice_code_pattern, text)
    invoice_number = re.findall(invoice_number_pattern, text)

    if invoice_code:
        invoice_code = invoice_code[0]
    else:
        invoice_code = ''

    if invoice_number:
        invoice_number = invoice_number[0]
    else:
        invoice_number = ''

    return {'发票代码': invoice_code, '发票号码': invoice_number}


def preprocess_image(image_path, max_size=4096, min_size=15, max_file_size=4 * 1024 * 1024):
    image = Image.open(image_path)
    # 调整图像尺寸
    width, height = image.size
    if width > max_size or height > max_size:
        ratio = min(max_size / width, max_size / height)
        new_width, new_height = int(width * ratio), int(height * ratio)
        image = image.resize((new_width, new_height), Image.ANTIALIAS)
    elif width < min_size or height < min_size:
        ratio = max(min_size / width, min_size / height)
        new_width, new_height = int(width * ratio), int(height * ratio)
        image = image.resize((new_width, new_height), Image.ANTIALIAS)

    # 检查文件大小
    image.save("temp.jpg", quality=95)
    file_size = os.path.getsize("temp.jpg")
    quality = 95
    while file_size > max_file_size and quality > 10:
        quality -= 5
        image.save("temp.jpg", quality=quality)
        file_size = os.path.getsize("temp.jpg")

    return "temp.jpg"


def get_vehicle_invoice_result(image_path, access_token):
    # 对图片进行预处理，压缩尺寸并调整图像质量
    enhanced_image = preprocess_image(image_path)

    # 发送 POST 请求到 OCR API
    with open(enhanced_image, 'rb') as f:
        img_data = f.read()

    # 使用百度 OCR API 进行识别
    response = requests.post(
        'https://aip.baidubce.com/rest/2.0/ocr/v1/vehicle_invoice?access_token={}'.format(access_token),
        headers={'Content-Type': 'application/x-www-form-urlencoded'},
        data={'image': base64.b64encode(img_data)},
    )

    # 延迟2秒钟
    time.sleep(2)

    if response.status_code != 200:
        raise ValueError('OCR API returned unexpected status code: {}'.format(response.status_code))

    # 从响应中提取识别结果
    json_data = response.json()
    if 'error_msg' in json_data:
        raise ValueError(json_data['error_msg'])

    # 解析识别结果，提取车辆发票信息
    words_result = json_data['words_result']
    invoice_fields = {}
    for field in words_result:
        if field['words'] in ['车辆类型', '厂牌型号', '车架号码', '发动机号码', '税率', '不含税金额', '税额',
                              '价税合计']:
            key = field['words']
            value = words_result[words_result.index(field) + 1]['words']
            invoice_fields[key] = value
        elif field['words'] in ['发票代码', '发票号码', '开票日期', '购买方名称', '销货单位名称']:
            key = field['words']
            value = field['value']
            invoice_fields[key] = value
    return invoice_fields
def get_ocr_result(image_path, access_token):
    try:
        url = "https://aip.baidubce.com/rest/2.0/ocr/v1/vehicle_license"
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        with open(image_path, 'rb') as f:
            img_data = f.read()

        img_base64 = base64.b64encode(img_data)
        data = {'image': img_base64}
        params = {'access_token': access_token}
        response = requests.post(url, headers=headers, data=data, params=params)
        result = response.json()
        logging.info("OCR Result:")
        logging.info(result)

        if 'error_code' in result:
            logging.warning('Error code: %s', result['error_code'])
            logging.warning('Error message: %s', result['error_msg'])
            return None

        return result['words_result']
    except Exception as e:
        logging.error('Error occurred while processing image %s', image_path)
        logging.error(e)
        return None

def clean_string(s):
    # 根据需要添加更多非法字符
    illegal_characters = ['\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', '\x08', '\x0B', '\x0C', '\x0E', '\x0F', '\x10', '\x11', '\x12', '\x13', '\x14', '\x15', '\x16', '\x17', '\x18', '\x19', '\x1A', '\x1B', '\x1C', '\x1D', '\x1E', '\x1F', '\x7F']
    for ch in illegal_characters:
        s = s.replace(ch, '')
    return s

def process_directory(directory_path, access_token, text_widget, root, is_invoice=False):
    # 创建 checkpoint.csv 文件（如果尚不存在）
    checkpoint_file = 'checkpoint.csv'
    if not os.path.exists(checkpoint_file):
        with open(checkpoint_file, 'w') as f:
            f.write('')

    # 读取已处理的子文件夹名字
    with open(checkpoint_file, 'r') as f:
        processed_subdirs = set(line.strip() for line in f)

    headers = ['子文件夹名称', '号牌号码', '车辆类型', '所有人', '住址', '发证单位', '使用性质', '品牌型号',
               '车辆识别代号', '发动机号码', '注册日期', '发证日期']
    headers_invoice = ['子文件夹名称', '开票日期', '机器编号', '购买方名称', '购买方身份证号码/组织机构代码', '车辆类型', '厂牌型号', '产地', '合格证号', '发动机号码', '车架号码', '价税合计', '价税合计小写', '销货单位名称', '销货单位电话', '销货单位纳税人识别号', '销货单位账号', '销货单位地址', '销货单位开户银行', '税率', '税额', '主管税务机关', '主管税务机关代码', '不含税价格', '限乘人数']

    # 创建一个新的 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.append(headers_invoice if is_invoice else headers)

    # 遍历文件夹下的所有子文件夹
    for subdir_name in os.listdir(directory_path):
        subdir_path = os.path.join(directory_path, subdir_name)
        if not os.path.isdir(subdir_path):
            continue

        # 如果子文件夹已经处理过，就跳过
        if subdir_name in processed_subdirs:
            logging.info(f'Skipping already processed directory {subdir_name}')
            continue

        # 处理当前子文件夹
        text_widget.insert(tk.END, f'正在处理子文件夹 {subdir_name}\n')
        text_widget.see(tk.END)
        root.update()

        # 遍历当前子文件夹下的所有图片
        for filename in os.listdir(subdir_path):
            if not filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                continue

            image_path = os.path.join(subdir_path, filename)
            logging.info(f'Processing image: {image_path}')

            # 根据 is_invoice 参数选择 OCR 函数
            if is_invoice:
                invoice_fields = get_vehicle_invoice_result(image_path, access_token)
                row_data = [subdir_name] + [clean_string(invoice_fields.get(key, '')) for key in headers_invoice[1:]]
            else:
                words_result = get_ocr_result(image_path, access_token)
                if words_result is not None:
                    row_data = [subdir_name] + [clean_string(words_result.get(key, {}).get('words', '')) for key in
                                                headers[1:]]
                else:
                    row_data = [subdir_name] + ['' for _ in headers[1:]]

            if row_data:
                # 将 row_data 追加到 Excel 文件中
                ws.append(row_data)
            else:
                logging.warning(f'Failed to process image: {image_path}')

        # 将处理完的子文件夹名字添加到 checkpoint.csv 文件中
        with open(checkpoint_file, 'a') as f:
            f.write(f'{subdir_name}\n')

    # 保存 Excel 文件
    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"
        try:
            wb.save(file_path)
            logging.info("Saving data to Excel file.")
            messagebox.showinfo("保存成功", "Excel 文件保存成功！")
            if messagebox.askyesno("打开文件夹", "Excel 文件保存成功，是否打开所在文件夹？"):
                directory = os.path.dirname(file_path)
                if sys.platform == "win32":
                    os.startfile(directory)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", directory])
                else:
                    subprocess.Popen(["xdg-open", directory])
        except Exception as e:
            logging.error("Failed to save data to Excel file: %s", e)
            messagebox.showerror("保存失败", "Excel 文件保存失败！")
    else:
        logging.warning(f'No data found. Skip saving data to Excel file.')
        messagebox.showwarning("无数据", "没有找到数据，Excel 文件未保存。")

def save_to_excel(data, headers, is_invoice=False):
    wb = Workbook()
    ws = wb.active

    # 设置字体
    font = Font(name='微软雅黑', size=12)

    # 设置边框
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # 设置底色
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 将表头写入Excel表格
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = font
        cell.border = border

    # 将数据写入Excel表格
    for row_idx, row_data in enumerate(data, start=2):
        ws.append(row_data)

        # 设置行高
        ws.row_dimensions[row_idx].height = 15

        # 根据奇偶行设置不同底色
        fill = fill1 if row_idx % 2 == 1 else fill2

        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font
            cell.border = border
            cell.fill = fill

            # 将“号牌号码”转换为大写
            if headers[col_idx - 1] == '号牌号码':
                cell.value = cell_value.upper()

            # 格式化“注册日期”和“发证日期”
            if headers[col_idx - 1] in ['注册日期', '发证日期']:
                try:
                    date_obj = datetime.strptime(cell_value, '%Y-%m-%d')
                    cell.number_format = 'yyyy/mm/dd'
                    cell.value = date_obj
                except ValueError:
                    pass

        logging.info(f"Writing row {row_idx}: {row_data}")

    return wb


def main():
    api_key = 'BlvkfH4d8XMTcyk2Lgd19KLy'
    secret_key = 'aodmQHoDZQIan78KXWshzXR71A112d51'

    access_token = get_access_token(api_key, secret_key)
    if access_token is None:
        logging.error("Unable to get access_token. Exiting the program.")
        messagebox.showerror("错误", "无法获取 access_token，程序退出。")
        return

    # 将以下变量声明为全局变量
    global wb
    global directory

    def process_directory_and_save_excel():
        # Debugging information
        global directory
        logging.info(f"process_directory_and_save_excel: Global directory is: {directory}")

        if directory:
            text.delete(1.0, tk.END)
            threading.Thread(target=process_directory, args=(directory, access_token, text, root), daemon=True).start()

    # 添加一个新的函数，用于处理机动车销售发票
    def process_directory_and_save_excel_invoice():
        global directory
        logging.info(f"process_directory_and_save_excel: Global directory is: {directory}")

        if directory:
            text.delete(1.0, tk.END)
            threading.Thread(target=process_directory, args=(directory, access_token, text, root, True),
                             daemon=True).start()

    def choose_directory():
        global directory
        directory = filedialog.askdirectory()
        if directory:
            directory_entry.delete(0, 'end')
            directory_entry.insert(0, directory)
            logging.info(f"Selected directory: {directory}")
        else:
            logging.warning("No directory selected.")

        # Debugging information
        logging.info(f"choose_directory: Global directory is now: {directory}")

    wb = None
    directory = None

    root = ThemedTk(theme="arc")
    root.title('技术质量部车辆信息提取器')

    # Create a frame to hold the buttons
    button_frame = ttk.Frame(root)
    button_frame.pack(padx=10, pady=10)

    browse_button = ttk.Button(button_frame, text='选择文件夹', command=choose_directory)
    browse_button.grid(row=0, column=0, padx=5, pady=5)

    directory_entry = ttk.Entry(button_frame, width=50)
    directory_entry.insert(0, '未选择文件夹')
    directory_entry.grid(row=0, column=1, padx=5, pady=5)

    start_button = ttk.Button(button_frame, text='开始匹配行驶证', command=process_directory_and_save_excel)
    start_button.grid(row=0, column=2, padx=5, pady=5)

    start_button_invoice = ttk.Button(button_frame, text='开始匹配机动车销售发票',
                                      command=process_directory_and_save_excel_invoice)
    start_button_invoice.grid(row=0, column=3, padx=5, pady=5)

    text = scrolledtext.ScrolledText(root, wrap='word', width=80, height=20)
    text.pack(padx=10, pady=10)

    # Set up logging to display in the text widget
    text_handler = TextHandler(text)
    logger = logging.getLogger()
    logger.addHandler(text_handler)

    root.mainloop()


if __name__ == '__main__':
    main()
